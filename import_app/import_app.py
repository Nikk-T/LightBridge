"""
LightBridge Import Tool — Standalone app (port 5001)
Drag-and-drop XLS/XLSX/CSV → maps.yaml converter.
All XLS parsing done server-side with openpyxl — no CDN required.
"""

from flask import Flask, render_template, jsonify, request
import yaml, shutil, io
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import csv as csvlib
    HAS_CSV = True
except ImportError:
    HAS_CSV = False

app = Flask(__name__, template_folder="templates", static_folder="static")

BASE_DIR      = Path(__file__).resolve().parent.parent
MAPS_PATH     = BASE_DIR / "config" / "maps.yaml"
SETTINGS_PATH = BASE_DIR / "config" / "settings.yaml"

# ── Config helpers ───────────────────────────────────────────
def load_maps() -> dict:
    if not MAPS_PATH.exists():
        return {}
    with open(MAPS_PATH) as f:
        return yaml.safe_load(f) or {}

def load_settings() -> dict:
    if not SETTINGS_PATH.exists():
        return {}
    with open(SETTINGS_PATH) as f:
        return yaml.safe_load(f) or {}

# ── Routes ───────────────────────────────────────────────────
@app.route("/")
def index():
    maps      = load_maps()
    maps_yaml = yaml.dump(maps, default_flow_style=False, sort_keys=False) if maps else ""
    return render_template("import.html", current_maps_yaml=maps_yaml)

@app.route("/api/parse_file", methods=["POST"])
def parse_file():
    """
    Accepts an uploaded XLS/XLSX/CSV file.
    Returns JSON: { sheets: { sheetName: [[row], [row], ...] } }
    """
    if "file" not in request.files:
        return jsonify({"status": "error", "message": "No file provided"})

    f    = request.files["file"]
    name = f.filename.lower()
    data = f.read()

    try:
        sheets = {}

        if name.endswith(".csv"):
            text = data.decode("utf-8-sig")
            reader = csvlib.reader(text.splitlines())
            rows = [list(row) for row in reader]
            # Normalise: replace empty strings with None
            rows = [[v if v != "" else None for v in row] for row in rows]
            sheets["Sheet1"] = rows

        elif name.endswith(".xlsx") or name.endswith(".xls"):
            wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
            for sheet_name in wb.sheetnames:
                ws   = wb[sheet_name]
                rows = []
                for row in ws.iter_rows(values_only=True):
                    rows.append([v for v in row])
                # Strip trailing all-None rows
                while rows and all(v is None for v in rows[-1]):
                    rows.pop()
                sheets[sheet_name] = rows
        else:
            return jsonify({"status": "error", "message": "Unsupported file type"})

        return jsonify({"status": "ok", "sheets": sheets})

    except Exception as e:
        return jsonify({"status": "error", "message": str(e)})

@app.route("/api/current_maps")
def current_maps():
    try:
        return jsonify({"status": "ok", "config": load_maps()})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)})

@app.route("/api/save_maps", methods=["POST"])
def save_maps():
    try:
        new_cfg = request.json.get("config", {})

        if not isinstance(new_cfg.get("unit_channel_map", {}), dict):
            return jsonify({"status": "error", "message": "unit_channel_map must be a dict"})
        if not isinstance(new_cfg.get("floor_channel_map", {}), dict):
            return jsonify({"status": "error", "message": "floor_channel_map must be a dict"})

        if MAPS_PATH.exists():
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            shutil.copy(MAPS_PATH, MAPS_PATH.parent / f"maps_backup_{ts}.yaml")

        MAPS_PATH.parent.mkdir(parents=True, exist_ok=True)
        with open(MAPS_PATH, "w") as f:
            yaml.dump(new_cfg, f, default_flow_style=False, sort_keys=False, allow_unicode=True)

        return jsonify({"status": "ok", "message": "maps.yaml saved"})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)})

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5001, debug=True)
